import os
import csv
from collections import defaultdict
import openpyxl
import argparse


# Ordine delle colonne nel dataset
COLUMNS = ['CG', 'ROC', 'NC', 'LC', 'IM', 'IQ', 'IdQ', 'LPQ']


def count_types_in_file(filepath):
    """Conta i code smell in un singolo file CSV o Excel"""
    counts = defaultdict(int)
    
    if filepath.lower().endswith('.csv'):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                if 'type' not in reader.fieldnames:
                    return counts
                for row in reader:
                    t = row['type'].strip().upper()
                    if t == "IDQ":
                        t = "IdQ"
                    if t in COLUMNS:
                        counts[t] += 1
        except:
            return counts
            
    elif filepath.lower().endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            for sheet in wb:
                header = next(sheet.iter_rows(values_only=True))
                header_lower = [str(h).lower() if h else '' for h in header]
                
                if "type" not in header_lower:
                    continue
                
                col_idx = header_lower.index("type")
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if col_idx < len(row) and row[col_idx]:
                        t = str(row[col_idx]).strip().upper()
                        if t == "IDQ":
                            t = "IdQ"
                        if t in COLUMNS:
                            counts[t] += 1
        except:
            return counts
    
    return counts


def process_repository(repo_path, repo_name):
    """Processa una repository: trova tutte le slice"""
    results = []
    
    # Trova tutte le cartelle slice in questa repository
    slice_folders = []
    for item in os.listdir(repo_path):
        item_path = os.path.join(repo_path, item)
        if os.path.isdir(item_path):
            slice_folders.append(item)
    
    # Ordina le cartelle per nome (così gli slice ID saranno in ordine)
    slice_folders.sort()
    
    # Processa ogni slice in ordine
    for i, slice_folder in enumerate(slice_folders, 1):
        slice_path = os.path.join(repo_path, slice_folder)
        slice_counts = defaultdict(int)
        
        # Conta tutti i file nella slice
        for file in os.listdir(slice_path):
            if file.lower().endswith(('.csv', '.xlsx', '.xls', '.xlsm', '.xlsb')):
                file_counts = count_types_in_file(os.path.join(slice_path, file))
                for k, v in file_counts.items():
                    slice_counts[k] += v
        
        # Crea riga per questa slice
        row = {
            'Repo': repo_name,
            'Slice ID': i,  # Numero progressivo (1, 2, 3...)
            'Anno': slice_folder,  # Nome della cartella = periodo temporale
        }
        
        # Aggiungi conteggi
        for col in COLUMNS:
            row[col] = slice_counts.get(col, 0)
        
        results.append(row)
    
    return results


def update_dataset(results_path, dataset_path):
    """Aggiorna il dataset Excel con i risultati"""
    all_data = []
    
    print(f"Analisi cartella: {results_path}")
    
    # Raccogli tutti i dati
    for item in sorted(os.listdir(results_path)):
        item_path = os.path.join(results_path, item)
        
        if os.path.isdir(item_path):
            repo_name = item
            print(f"\nRepository: {repo_name}")
            
            repo_data = process_repository(item_path, repo_name)
            
            if repo_data:
                print(f"  Trovate {len(repo_data)} slice")
                for data in repo_data:
                    print(f"    Slice {data['Slice ID']} ({data['Anno']}): {data}")
                all_data.extend(repo_data)
            else:
                print(f"  Nessuna slice trovata")
    
    if not all_data:
        print("\nNessun dato trovato!")
        return False
    
    print(f"\nTotale slice trovate: {len(all_data)}")
    
    # Carica o crea Excel
    if os.path.exists(dataset_path):
        wb = openpyxl.load_workbook(dataset_path)
        if "dataset" in wb.sheetnames:
            ws = wb["dataset"]
        else:
            ws = wb.active
            ws.title = "dataset"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "dataset"
    
    # Intestazione
    headers = ['Repo', 'Slice ID', 'Anno', 'CG', 'ROC', 'NC', 'LC', 'IM', 'IQ', 'IdQ', 'LPQ']
    
    # Cancella dati esistenti (tranne intestazione)
    if ws.max_row > 1:
        existing_headers = []
        for i in range(1, len(headers) + 1):
            existing_headers.append(ws.cell(row=1, column=i).value)
        
        if existing_headers == headers:
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)
        else:
            ws.delete_rows(1, ws.max_row)
            ws.append(headers)
    else:
        ws.append(headers)
    
    # Stile intestazione
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Scrivi tutti i dati
    for data in all_data:
        row = [
            data['Repo'],
            data['Slice ID'],
            data['Anno'],
            data['CG'],
            data['ROC'],
            data['NC'],
            data['LC'],
            data['IM'],
            data['IQ'],
            data['IdQ'],
            data['LPQ']
        ]
        ws.append(row)
    
    # Aggiusta larghezza colonne
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    # Salva
    wb.save(dataset_path)
    print(f"\nDataset salvato: {dataset_path}")
    print(f"Righe scritte: {len(all_data)}")
    
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dataset', '-d', required=True, help='Percorso file dataset Excel')
    args = parser.parse_args()
    
    # Cartella results
    script_dir = os.path.dirname(os.path.abspath(__file__))
    results_path = os.path.join(script_dir, '..', 'results')
    results_path = os.path.normpath(results_path)
    
    if not os.path.exists(results_path):
        print(f"ERRORE: Cartella 'results' non trovata in {results_path}")
        return
    
    print(f"Cartella results: {results_path}")
    print(f"File dataset: {args.dataset}")
    print("-" * 50)
    
    success = update_dataset(results_path, args.dataset)
    
    if success:
        print("\nOperazione completata con successo!")
    else:
        print("\nOperazione fallita.")


if __name__ == "__main__":
    main()