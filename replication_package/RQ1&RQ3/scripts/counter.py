import os
import argparse
import csv
from collections import defaultdict
import openpyxl


COLUMN_ORDER = ['CG', 'ROC', 'NC', 'LC', 'IM', 'IdQ', 'IQ', 'LPQ']


def count_types_in_csv(filepath):
    type_counts = defaultdict(int)
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            if 'type' not in reader.fieldnames:
                return {}, False
            for row in reader:
                t = row['type'].strip().upper()
                if t == "IDQ":
                    t = "IdQ"
                if t in COLUMN_ORDER:
                    type_counts[t] += 1
        return dict(type_counts), True
    except:
        return {}, False


def count_types_in_excel(filepath):
    type_counts = defaultdict(int)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        for sheet in wb:
            try:
                header = next(sheet.iter_rows(values_only=True))
                if "type" not in header:
                    continue
                col_idx = header.index("type") + 1
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= col_idx:
                        t = str(row[col_idx-1]).strip().upper()
                        if t == "IDQ":
                            t = "IdQ"
                        if t in COLUMN_ORDER:
                            type_counts[t] += 1
            except:
                continue
        return dict(type_counts), True
    except:
        return {}, False


def process_folder(folder_path):
    exts = ('.xlsx', '.xls', '.xlsm', '.xlsb', '.csv')
    files = [f for f in os.listdir(folder_path) if f.lower().endswith(exts)]
    if not files:
        return None
    total_counts = defaultdict(int)
    for f in files:
        fp = os.path.join(folder_path, f)
        if f.lower().endswith('.csv'):
            counts, ok = count_types_in_csv(fp)
        else:
            counts, ok = count_types_in_excel(fp)
        if ok and counts:
            for k, v in counts.items():
                total_counts[k] += v
    return {t: total_counts.get(t, 0) for t in COLUMN_ORDER} if total_counts else None


def update_dataset(results_path, dataset_path):
    try:
        if os.path.exists(dataset_path):
            wb = openpyxl.load_workbook(dataset_path)
            for name in wb.sheetnames:
                if name.lower() == "dataset":
                    ws = wb[name]
                    break
            else:
                ws = wb.create_sheet("dataset")
                ws.append(['Repo', 'CG', 'ROC', 'NC', 'LC', 'IM', 'IQ', 'IdQ', 'LPQ'])
                for cell in ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            ws = wb.create_sheet("dataset")
            ws.append(['Repo', 'CG', 'ROC', 'NC', 'LC', 'IM', 'IQ', 'IdQ', 'LPQ'])
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)

        existing = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), 2):
            if row[0]:
                existing[row[0]] = i

        processed = updated = added = 0
        for item in os.listdir(results_path):
            full = os.path.join(results_path, item)
            if not os.path.isdir(full):
                continue
            repo = item.replace('\\', '/')
            counts = process_folder(full)
            if not counts:
                continue
            processed += 1
            if repo in existing:
                row = existing[repo]
                ws.cell(row=row, column=2, value=counts.get('CG', 0))
                ws.cell(row=row, column=3, value=counts.get('ROC', 0))
                ws.cell(row=row, column=4, value=counts.get('NC', 0))
                ws.cell(row=row, column=5, value=counts.get('LC', 0))
                ws.cell(row=row, column=6, value=counts.get('IM', 0))
                ws.cell(row=row, column=7, value=counts.get('IQ', 0))
                ws.cell(row=row, column=8, value=counts.get('IdQ', 0))
                ws.cell(row=row, column=9, value=counts.get('LPQ', 0))
                updated += 1
            else:
                ws.append([
                    repo,
                    counts.get('CG', 0),
                    counts.get('ROC', 0),
                    counts.get('NC', 0),
                    counts.get('LC', 0),
                    counts.get('IM', 0),
                    counts.get('IQ', 0),
                    counts.get('IdQ', 0),
                    counts.get('LPQ', 0)
                ])
                added += 1

        for col in ws.columns:
            ml = 0
            for cell in col:
                try:
                    ml = max(ml, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 50)

        wb.save(dataset_path)
        print(f"Processate: {processed}, Aggiornate: {updated}, Aggiunte: {added}")
        return True
    except Exception as e:
        print(f"Errore: {e}")
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dataset', '-d', required=True)
    args = parser.parse_args()
    
    results_path = os.path.normpath(os.path.join(os.getcwd(), "..", "results"))
    if not os.path.isdir(results_path):
        print(f"Cartella non trovata: {results_path}")
        return
    
    update_dataset(results_path, args.dataset)


if __name__ == "__main__":
    main()